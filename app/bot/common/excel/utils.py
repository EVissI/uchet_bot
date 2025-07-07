import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime

from app.bot.common.texts import get_text
from app.db.models import (
    Check,
    Object,
    ObjectCheck,
    ObjectProficAccounting,
    ProficAccounting,
    Tool,
    User,
)


def create_expense_report(expenses: list[ObjectCheck], lang: str) -> io.BytesIO:
    """
    Creates an Excel report from expenses list.

    Args:
        expenses: List of ObjectCheck instances
        lang: Language code for translations

    Returns:
        io.BytesIO: Excel file as bytes buffer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("expense_sheet_name", lang)

    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    # Headers
    headers = [
        "ID",
        get_text("date_column", lang),
        get_text("description_column", lang),
        get_text("amount_column", lang),
        get_text("expense_type_column", lang),
        get_text("user_column", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data
    total_amount = 0
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.id)
        ws.cell(row=row, column=2, value=expense.created_at.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=3, value=expense.description)
        ws.cell(row=row, column=4, value=expense.amount)
        ws.cell(
            row=row,
            column=5,
            value=(
                get_text("own_expense", lang)
                if expense.own_expense
                else get_text("company_expense", lang)
            ),
        )
        ws.cell(row=row, column=6, value=expense.user.user_enter_fio)

        total_amount += expense.amount

    # Total row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=3, value=get_text("total_amount", lang))
    total_cell = ws.cell(row=total_row, column=4, value=total_amount)
    total_cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 10  # ID
    ws.column_dimensions["B"].width = 15  # Date
    ws.column_dimensions["C"].width = 40  # Description
    ws.column_dimensions["D"].width = 15  # Amount
    ws.column_dimensions["E"].width = 20  # Expense type
    ws.column_dimensions["F"].width = 30  # User

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def create_tools_report(tools: list[Tool], lang: str) -> io.BytesIO:
    """Create Excel report for tools list"""
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("excel_tools_sheet_name", lang)

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    headers = [
        get_text("excel_tool_name", lang),
        get_text("excel_tool_description", lang),
        get_text("excel_tool_status", lang),
        get_text("excel_tool_user", lang),
        get_text("excel_tool_date", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data
    for row, tool in enumerate(tools, 2):
        ws.cell(row=row, column=1, value=tool.name)
        ws.cell(row=row, column=2, value=tool.description)
        ws.cell(row=row, column=3, value=tool.status)
        ws.cell(row=row, column=4, value=tool.user.user_enter_fio if tool.user else "-")
        ws.cell(row=row, column=5, value=tool.created_at.strftime("%d.%m.%Y"))

    # Column widths
    ws.column_dimensions["A"].width = 30  # Name
    ws.column_dimensions["B"].width = 40  # Description
    ws.column_dimensions["C"].width = 20  # Status
    ws.column_dimensions["D"].width = 30  # User
    ws.column_dimensions["E"].width = 15  # Date

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def create_user_tools_report(tools: list[Tool], user: User, lang: str) -> io.BytesIO:
    """
    Формирует Excel-файл со списком инструментов пользователя.
    :param tools: список инструментов (Tool)
    :param user: объект пользователя (User)
    :param lang: язык для переводов
    :return: io.BytesIO с Excel-файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Инструменты {user.user_enter_fio}"[:31]

    # Стили
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    # Заголовки
    headers = [
        get_text("excel_tool_name", lang),
        get_text("excel_tool_description", lang),
        get_text("excel_tool_status", lang),
        get_text("excel_tool_date", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Данные
    for row, tool in enumerate(tools, 2):
        ws.cell(row=row, column=1, value=tool.name)
        ws.cell(row=row, column=2, value=tool.description)
        ws.cell(row=row, column=3, value=tool.status)
        ws.cell(
            row=row,
            column=4,
            value=(
                tool.created_at.strftime("%d.%m.%Y")
                if hasattr(tool, "created_at") and tool.created_at
                else ""
            ),
        )

    # Ширина колонок
    ws.column_dimensions["A"].width = 30  # Name
    ws.column_dimensions["B"].width = 40  # Description
    ws.column_dimensions["C"].width = 20  # Status
    ws.column_dimensions["D"].width = 15  # Date

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def generate_transfer_template(lang: str) -> io.BytesIO:
    """
    Generate Excel template for bulk tools transfer

    Args:
        lang: Language code for translations

    Returns:
        io.BytesIO: Excel template file as bytes buffer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("transfer_template_sheet", lang)

    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    headers = [
        get_text("excel_tool_id", lang),
        get_text("excel_recipient_username", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    example_row = ["12345", "@username или 123456789"]
    for col, value in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = value
        cell.font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 15  # Tool ID
    ws.column_dimensions["B"].width = 35  # Recipient username/id

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def create_tools_export(tools: list[Tool], lang: str) -> io.BytesIO:
    """Create Excel file with tools list"""
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("tools_export_sheet", lang)

    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    # Headers
    headers = [
        get_text("excel_tool_id", lang),
        get_text("excel_tool_name", lang),
        get_text("excel_tool_status", lang),
        get_text("excel_tool_owner", lang),
        get_text("excel_tool_description", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data
    for row, tool in enumerate(tools, 2):
        ws.cell(row=row, column=1, value=tool.id)
        ws.cell(row=row, column=2, value=tool.name)
        ws.cell(row=row, column=3, value=tool.status)
        ws.cell(row=row, column=4, value=tool.user.user_enter_fio if tool.user else "-")
        ws.cell(row=row, column=5, value=tool.description or "-")

    # Column widths
    ws.column_dimensions["A"].width = 10  # ID
    ws.column_dimensions["B"].width = 30  # Name
    ws.column_dimensions["C"].width = 15  # Status
    ws.column_dimensions["D"].width = 30  # Owner
    ws.column_dimensions["E"].width = 40  # Description

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def create_materials_excel(materials: list, lang: str) -> io.BytesIO:
    """Create Excel file with materials list"""
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("materials_sheet_name", lang)

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        get_text("excel_material_id", lang),
        get_text("excel_material_description", lang),
        get_text("excel_material_location", lang),
        get_text("excel_material_file_id", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data
    for row, material in enumerate(materials, 2):
        ws.cell(row=row, column=1, value=material.id)
        ws.cell(row=row, column=2, value=material.description)
        ws.cell(row=row, column=3, value=material.storage_location or "-")
        ws.cell(row=row, column=4, value=material.file_id or "-")

    # Adjust column widths
    ws.column_dimensions["A"].width = 10  # ID
    ws.column_dimensions["B"].width = 50  # Description
    ws.column_dimensions["C"].width = 30  # Location
    ws.column_dimensions["D"].width = 70  # File ID

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def generate_tmc_template(lang: str = "ru") -> io.BytesIO:
    """
    Generate Excel template for bulk tools upload

    Args:
        lang: Language code for translations

    Returns:
        io.BytesIO: Excel template file as bytes buffer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("tmc_template_sheet", lang)

    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    # Headers
    headers = [
        get_text("excel_tool_name", lang),
        get_text("excel_tool_quantity", lang),
        get_text("excel_tool_description", lang),
        get_text("excel_tool_status", lang),
        get_text("excel_tool_file_id", lang),
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Example row
    example_row = [
        "Перфоратор Bosch GBH 2-26",
        "2",
        "Новый, в комплекте 2 бура",
        "свободный",
        "AgACAgIAAxkBAAIBo2VrR5B...",
    ]
    for col, value in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = value
        cell.font = Font(italic=True, color="808080")

    # Column widths
    ws.column_dimensions["A"].width = 40  # Name
    ws.column_dimensions["B"].width = 15  # Quantity
    ws.column_dimensions["C"].width = 50  # Description
    ws.column_dimensions["D"].width = 15  # Status
    ws.column_dimensions["E"].width = 70  # File ID

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def create_object_financial_report(
    object_id: int,
    start_date: datetime,
    end_date: datetime,
    lang: str,
    profic_records: list[ProficAccounting],
    object_checks: list[ObjectCheck],
    object: Object,
) -> io.BytesIO:
    """Create financial report for specific object"""
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("object_finance_sheet", lang)

    ws.merge_cells("A1:F1")
    header_cell = ws.cell(row=1, column=1, value=f"🏗 {object.name}")
    header_cell.font = Font(bold=True, size=14)

    ws.cell(
        row=2,
        column=1,
        value=f"📅 {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
    )

    return _process_financial_data(
        ws=ws,
        profic_records=profic_records,
        checks=object_checks,
        lang=lang,
        start_row=4,
    )


def create_outobject_financial_report(
    start_date: datetime,
    end_date: datetime,
    lang: str,
    checks: list[Check],
) -> io.BytesIO:
    """Create financial report for expenses outside objects"""
    wb = Workbook()
    ws = wb.active
    ws.title = get_text("outobject_finance_sheet", lang)

    # Add report header
    ws.merge_cells("A1:F1")
    header_cell = ws.cell(
        row=1, column=1, value=get_text("outobject_expenses_header", lang)
    )
    header_cell.font = Font(bold=True, size=14)

    ws.cell(
        row=2,
        column=1,
        value=f"📅 {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
    )

    # Add headers and process data
    return _process_financial_data(
        ws=ws,
        checks=checks,
        profic_records=[],  # No profic records for out-object report
        lang=lang,
        start_row=4,
    )


def _process_financial_data(
    ws: Worksheet,
    profic_records: list[ProficAccounting],
    checks: list[Check | ObjectCheck],
    lang: str,
    start_row: int = 1,
) -> io.BytesIO:
    """Helper function to process financial data and create report"""
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    income_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    expense_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    # Headers
    headers = [
        get_text("finance_date", lang),
        get_text("finance_type", lang),
        get_text("finance_category", lang),
        get_text("finance_description", lang),
        get_text("finance_amount", lang),
        get_text("finance_user", lang),
        "file_id"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    current_row = start_row + 1
    total_income = 0
    total_expense = 0

    # Process profit/loss records
    for record in profic_records:
        row = [
            record.created_at.strftime("%d.%m.%Y"),
            record.payment_type.value,
            get_text("direct_payment", lang),
            record.purpose,
            record.amount,
            record.user.user_enter_fio,
            getattr(check, "file_id", ""), 
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.fill = (
                income_fill
                if record.payment_type == ProficAccounting.PaymentType.income
                else expense_fill
            )

        if record.payment_type == ProficAccounting.PaymentType.income:
            total_income += record.amount
        else:
            total_expense += record.amount
        current_row += 1

    # Process checks
    for check in checks:
        row = [
            check.created_at.strftime("%d.%m.%Y"),
            get_text("expense", lang),
            get_text("check_payment", lang),
            check.description,
            check.amount,
            check.user.user_enter_fio,
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.fill = expense_fill
        total_expense += check.amount
        current_row += 1

    # Add summary
    summary_row = current_row + 2
    _add_summary_section(
        ws=ws,
        row=summary_row,
        total_income=total_income,
        total_expense=total_expense,
        lang=lang,
    )

    # Set column widths
    _set_column_widths(ws)

    # Save to buffer
    excel_buffer = io.BytesIO()
    ws.parent.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def _add_summary_section(
    ws: Worksheet, row: int, total_income: float, total_expense: float, lang: str
) -> None:
    """Add summary section to worksheet"""
    ws.cell(row=row, column=1, value=get_text("total_income", lang))
    ws.cell(row=row, column=5, value=total_income).font = Font(bold=True)

    ws.cell(row=row + 1, column=1, value=get_text("total_expense", lang))
    ws.cell(row=row + 1, column=5, value=total_expense).font = Font(bold=True)

    profit = total_income - total_expense
    profit_cell = ws.cell(row=row + 2, column=1, value=get_text("total_profit", lang))
    amount_cell = ws.cell(row=row + 2, column=5, value=profit)
    profit_cell.font = Font(bold=True)
    amount_cell.font = Font(bold=True)
    if profit < 0:
        amount_cell.fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )


def _set_column_widths(ws: Worksheet) -> None:
    """Set column widths for worksheet"""
    widths = {
        "A": 15,  # Date
        "B": 15,  # Type
        "C": 20,  # Category
        "D": 50,  # Description
        "E": 15,  # Amount
        "F": 30,  # User
        "G": 50,  # file_id
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_profic_report(
    profic_records: list[ProficAccounting],
    object_checks: list[ObjectCheck] | None = None,
    non_object_checks: list[Check] | None = None,
    start_date: datetime = None,
    end_date: datetime = None,
    lang: str = "ru",
) -> io.BytesIO:
    """Creates financial report in Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансовый отчет"

    # Styles
    header_style = {
        "font": Font(bold=True, size=12),
        "fill": PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }

    income_style = {
        "fill": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

    expense_style = {
        "fill": PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

    # Report header
    ws.merge_cells("A1:G1")
    header_cell = ws["A1"]
    header_cell.value = "ФИНАНСОВЫЙ ОТЧЕТ"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal="center")

    # Period info
    if start_date and end_date:
        ws.merge_cells("A2:G2")
        period_cell = ws["A2"]
        period_cell.value = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        period_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Дата",
        "Тип",
        "Объект",
        "Описание",
        "Сумма",
        "За свой счет",
        "Пользователь",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        for style_attr, style_value in header_style.items():
            setattr(cell, style_attr, style_value)

    # Data
    current_row = 5
    total_income = 0
    total_expense = 0
    for record in profic_records:
        if hasattr(record, "object") and getattr(record, "object", None):
            object_name = record.object.name
        else:
            object_name = "-"
        row_data = [
            record.created_at.strftime("%d.%m.%Y"),
            (
                record.payment_type.value
                if hasattr(record.payment_type, "value")
                else record.payment_type
            ),
            object_name,
            record.purpose,
            record.amount,
            "-",
            record.user.user_enter_fio,
        ]
        style = (
            income_style
            if (
                getattr(record, "payment_type", None)
                == ProficAccounting.PaymentType.income
                or getattr(record, "payment_type", None)
                == ProficAccounting.PaymentType.income.value
            )
            else expense_style
        )

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            for style_attr, style_value in style.items():
                setattr(cell, style_attr, style_value)

        if (
            getattr(record, "payment_type", None) == ProficAccounting.PaymentType.income
            or getattr(record, "payment_type", None)
            == ProficAccounting.PaymentType.income.value
        ):
            total_income += record.amount
        else:
            total_expense += record.amount

        current_row += 1

    # Process object checks
    if object_checks:
        for check in object_checks:
            row_data = [
                check.created_at.strftime("%d.%m.%Y"),
                "Чек (объект)",
                check.object.name,
                check.description or "-",
                check.amount,
                "Да" if check.own_expense else "Нет",
                check.user.user_enter_fio,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                for style_attr, style_value in expense_style.items():
                    setattr(cell, style_attr, style_value)

            total_expense += check.amount
            current_row += 1

    # Process personal checks
    if non_object_checks:
        for check in non_object_checks:
            row_data = [
                check.created_at.strftime("%d.%m.%Y"),
                "Чек (вне объекта)",  # Changed label
                "-",
                check.description or "-",
                check.amount,
                "Да" if check.own_expense else "Нет",
                check.user.user_enter_fio,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                for style_attr, style_value in expense_style.items():
                    setattr(cell, style_attr, style_value)

            total_expense += check.amount
            current_row += 1

    # Totals
    summary_row = current_row + 2
    summary_style = {"font": Font(bold=True)}

    ws.cell(row=summary_row, column=1, value="ИТОГО доходы:").font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=total_income).font = Font(bold=True)

    ws.cell(row=summary_row + 1, column=1, value="ИТОГО расходы:").font = Font(
        bold=True
    )
    ws.cell(row=summary_row + 1, column=5, value=total_expense).font = Font(bold=True)

    profit = total_income - total_expense
    ws.cell(row=summary_row + 2, column=1, value="ИТОГО прибыль:").font = Font(
        bold=True
    )
    profit_cell = ws.cell(row=summary_row + 2, column=5, value=profit)
    profit_cell.font = Font(bold=True)
    if profit < 0:
        profit_cell.fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

    # Column widths
    column_widths = [15, 15, 30, 50, 15, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer

def create_workers_full_info_excel(users: list[User]) -> io.BytesIO:
    """
    Формирует Excel-файл со всеми рабочими:
    - ФИО, Telegram ID, username, телефон, роль
    - Список объектов (названия через запятую)
    - Список инструментов (id:имя через запятую)
    - Список документов (file_id через запятую)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Рабочие"

    headers = [
        "ФИО", "Telegram ID", "Username", "Телефон", "Роль",
        "Объекты", "Инструменты (id:имя)", "Документы (file_id)"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for user in users:
        fio = user.user_enter_fio
        telegram_id = user.telegram_id
        username = user.username
        phone = user.phone_number
        role = user.role.value if hasattr(user.role, "value") else user.role

        objects = ", ".join(obj.name for obj in user.objects) if user.objects else ""
        tools = " ; ".join(f"{tool.id}:{tool.name}" for tool in user.tools) if user.tools else ""
        documents = " , ".join(doc.file_id for doc in user.documents) if user.documents else ""

        ws.append([fio, telegram_id, username, phone, role, objects, tools, documents])

    # Автоширина
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(15, min(max_length + 2, 50))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
