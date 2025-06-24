from aiogram import Router, F
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from loguru import logger
from openpyxl import load_workbook
import io

from app.bot.common.excel.utils import generate_tmc_template
from app.bot.common.fsm_managment import DialogMessageManager, StateHistoryMixin
from app.bot.common.states import AdminPanelStates, TMCStates
from app.bot.common.texts import get_all_texts, get_text
from app.bot.filters.user_info import UserInfo
from app.bot.kbds.inline_kbds import TMCCallback, build_tmc_upload_kbd
from app.bot.kbds.markup_kbds import stop_kb
from app.db.models import Tool, User
from app.db.database import async_session_maker
from app.db.dao import ToolDAO
from app.db.schemas import ToolModel
from app.bot.kbds.markup_kbds import AdminToolsControlKeyboard


class TMCRouter(Router, StateHistoryMixin, DialogMessageManager):
    def __init__(self):
        super().__init__()


tmc_router = TMCRouter()


@tmc_router.message(
    F.text.in_(get_all_texts("tmc_upload_btn")),
    StateFilter(AdminPanelStates.tools_control),
    UserInfo(),
)
async def process_tmc_upload_btn(message: Message, user_info: User):
    """Handle TMC upload button click"""
    logger.info(f"User {user_info.telegram_id} started TMC upload process")
    await message.answer(
        text=get_text("tmc_upload_start", user_info.language),
        reply_markup=build_tmc_upload_kbd(user_info.language),
    )

@tmc_router.message(F.text.in_(get_all_texts("stop_upload_btn")), StateFilter(TMCStates),UserInfo())
async def process_tmc_stop(message: Message, state: FSMContext, user_info: User):
    logger.info(f"User {user_info.telegram_id} stopped TMC upload process")
    await state.set_state(AdminPanelStates.tools_control)
    await message.answer(
        text=get_text("tmc_upload_stopped", user_info.language),
        reply_markup=AdminToolsControlKeyboard.build_tools_control_kb(
            user_info.language
        ),
    )
    await tmc_router.clear_messages(state, message.chat.id, message.bot)

@tmc_router.callback_query(TMCCallback.filter(F.action == "manual"),UserInfo())
async def process_tmc_manual(
    callback: CallbackQuery, state: FSMContext, user_info: User
):
    """Start manual TMC input process"""
    logger.info(f"User {user_info.telegram_id} selected manual TMC input")
    await state.set_state(TMCStates.input_name)
    await callback.message.delete()
    await callback.message.answer(
        text=get_text(
            "tmc_enter_name",
            user_info.language,
        ),reply_markup=stop_kb(user_info.language),
    )
    await callback.answer()


@tmc_router.message(F.text, StateFilter(TMCStates.input_name),UserInfo())
async def process_tmc_name(message: Message, state: FSMContext, user_info: User):
    logger.info(f"User {user_info.telegram_id} entered TMC name: {message.text}")
    await tmc_router.save_message(state, message.message_id)
    await state.update_data(name=message.text)
    await state.set_state(TMCStates.input_description)
    bot_message = await message.answer(
        get_text("tmc_enter_description", user_info.language)
    )
    await tmc_router.save_message(state, bot_message.message_id)


@tmc_router.message(F.text, StateFilter(TMCStates.input_description),UserInfo())
async def process_tmc_unit(message: Message, state: FSMContext, user_info: User):
    await tmc_router.save_message(state, message.message_id)
    await state.update_data(description=message.text)
    await state.set_state(TMCStates.input_file)
    bot_message = await message.answer(
        get_text("tmc_enter_file", user_info.language)
    )
    await tmc_router.save_message(state, bot_message.message_id)


@tmc_router.message(F.photo, StateFilter(TMCStates.input_file),UserInfo())
async def process_tmc_file(message: Message, state: FSMContext, user_info: User):
    logger.info(f"User {user_info.telegram_id} uploaded TMC file")
    try:
        await tmc_router.save_message(state, message.message_id)
        if not message.photo:
            bot_messsage = await message.answer(
                get_text("tmc_invalid_file", user_info.language)
            )
            await tmc_router.save_message(state, bot_messsage.message_id)
            return
        file_id = message.photo[-1].file_id
        await state.update_data(file_id=file_id)

        await state.set_state(TMCStates.input_quantity)
        bot_messsage = await message.answer(
            get_text("tmc_enter_quantity", user_info.language)
        )
        await tmc_router.save_message(state, bot_messsage.message_id)

    except Exception as e:
        logger.error(
            f"Error processing TMC file for user {user_info.telegram_id}: {str(e)}"
        )
        await message.answer(get_text("tmc_file_error", user_info.language))


@tmc_router.message(F.text, StateFilter(TMCStates.input_quantity),UserInfo())
async def process_tmc_quantity(message: Message, state: FSMContext, user_info: User):
    try:
        await tmc_router.save_message(state, message.message_id)
        quantity = int(message.text)
        if quantity <= 0:
            raise ValueError("Quantity must be positive")
        data = await state.get_data()
        async with async_session_maker() as session:
            tools = []
            for _ in range(quantity):
                tools.append(
                    ToolModel(
                        name=data.get("name"),
                        description=data.get("description"),
                        file_id=data.get("file_id"),
                        status=Tool.Status.free,
                    )
                )
            await ToolDAO.add_many(session, tools)
        logger.info(f"User {user_info.telegram_id} added {quantity} TMC items")
        bot_message = await message.answer(
            get_text(
                "tmc_upload_complete",
                user_info.language,
                name=data.get("name"),
                count=quantity,
            )
        )
        await tmc_router.save_message(state, bot_message.message_id)
        await state.set_state(TMCStates.input_name)
        bot_message = await message.answer(
            get_text("tmc_enter_name", user_info.language)
        )
        await tmc_router.save_message(state, bot_message.message_id)
    except ValueError as e:
        logger.error(str(e))
        bot_message = await message.answer(
            get_text("tmc_invalid_quantity", user_info.language)
        )
        await tmc_router.save_message(state, bot_message.message_id)



@tmc_router.callback_query(TMCCallback.filter(F.action == "template"),UserInfo())
async def process_tmc_template(callback: CallbackQuery, user_info: User):
    """Handle TMC template download"""
    logger.info(f"User {user_info.telegram_id} requested TMC template")
    try:
        template = generate_tmc_template(user_info.language)
        await callback.message.answer_document(
            document=BufferedInputFile(
                template.getvalue(), filename="tmc_upload_template.xlsx"
            ),
            caption=get_text("tmc_template_instruction", user_info.language),
        )
        await callback.answer()
    except Exception as e:
        logger.error(
            f"Error generating TMC template for user {user_info.telegram_id}: {str(e)}"
        )
        await callback.answer(
            get_text("template_error", user_info.language), show_alert=True
        )


@tmc_router.callback_query(TMCCallback.filter(F.action == "upload"),UserInfo())
async def process_tmc_upload_start(
    callback: CallbackQuery, state: FSMContext, user_info: User
):
    """Handle TMC file upload start"""
    logger.info(f"User {user_info.telegram_id} started TMC file upload process")
    try:
        await callback.message.delete()
        await callback.message.answer(
            text=get_text("tmc_upload_instruction", user_info.language),reply_markup=stop_kb(user_info.language)
        )
        await state.set_state(TMCStates.input_file)
        await callback.answer()
        logger.debug(f"State set to waiting_file for user {user_info.telegram_id}")
    except Exception as e:
        logger.error(
            f"Error starting TMC upload for user {user_info.telegram_id}: {str(e)}"
        )


@tmc_router.message(F.document, StateFilter(TMCStates.input_file),UserInfo())
async def process_tmc_file_upload(message: Message, state: FSMContext, user_info: User):
    """Process uploaded TMC Excel file"""
    logger.info(f"User {user_info.telegram_id} uploaded TMC Excel file")
    try:
        if (
            not message.document.mime_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            await message.answer(get_text("invalid_file_type", user_info.language))
            return

        # Download file
        file = await message.bot.get_file(message.document.file_id)
        file_content = await message.bot.download_file(file.file_path)

        # Load workbook
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active

        # Parse tools data
        tools_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue

            name, quantity, description, status, file_id = row

            # Validate required fields
            if not name or not quantity:
                continue

            try:
                quantity = int(quantity)
                if quantity <= 0:
                    continue
            except (ValueError, TypeError):
                continue

            tools_data.append(
                {
                    "name": str(name),
                    "quantity": quantity,
                    "description": str(description) if description else None,
                    "status": Tool.Status.free if not status else status.lower(),
                    "file_id": str(file_id) if file_id else None,
                }
            )

        if not tools_data:
            await message.answer(get_text("no_valid_tools", user_info.language))
            return

        # Save tools to database
        success_count = 0
        error_count = 0

        async with async_session_maker() as session:
            for tool_data in tools_data:
                try:
                    quantity = tool_data.pop("quantity")
                    tools = []
                    for _ in range(quantity):
                        tools.append(ToolModel(**tool_data))

                    await ToolDAO.add_many(session, tools)
                    success_count += quantity
                except Exception as e:
                    logger.error(f"Error saving tool {tool_data}: {e}")
                    error_count += 1

        await message.answer(
            get_text(
                "tmc_bulk_upload_complete",
                user_info.language,
                success=success_count,
                errors=error_count,
            )
        )

        await state.set_state(AdminPanelStates.tools_control)
        await message.answer(
            text=get_text("return_to_tools_control", user_info.language),
            reply_markup=AdminToolsControlKeyboard.build_tools_control_kb(
                user_info.language
            ),
        )

    except Exception as e:
        logger.error(
            f"Error processing TMC file from user {user_info.telegram_id}: {e}"
        )
        await message.answer(get_text("tmc_file_error", user_info.language))
