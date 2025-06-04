import random
from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from app.bot.common.texts import get_text
from app.bot.filters.user_info import UserInfo
from app.db.dao import (
    ObjectCheckDAO,
    ObjectDAO,
    ObjectDocumentDAO,
    ObjectMemberDAO,
    ToolDAO,
)
from app.db.database import async_session_maker
from app.db.schemas import (
    ObjectCheckModel,
    ObjectDocumentModel,
    ObjectFilterModel,
    ObjectMemberModel,
    ObjectModel,
    ToolModel,

)
from app.config import settings
from app.db.models import ObjectDocument, User, Tool

admin_mock_router = Router()

SAMPLE_TOOLS = [
    ("Дрель", "Аккумуляторная дрель Makita"),
    ("Перфоратор", "Перфоратор Bosch GBH 2-26"),
    ("Болгарка", "УШМ Metabo WE 15-125"),
    ("Шуруповерт", "Шуруповерт DeWalt DCD791"),
    ("Лобзик", "Электролобзик Festool PS 420"),
]

SAMPLE_OBJECTS = [
    ("ЖК Солнечный", "Строительство жилого комплекса"),
    ("Бизнес-центр Престиж", "Реконструкция офисного здания"),
    ("Школа №5", "Капитальный ремонт"),
    ("Торговый центр Заря", "Отделочные работы"),
    ("Склад Логистик", "Строительство складского помещения"),
]


class MockToolsStates(StatesGroup):
    waiting_photo = State()
    waiting_quantity = State()



class MockObjectChecksStates(StatesGroup):
    waiting_photo = State()
    waiting_quantity = State()

class MockDocumentsStates(StatesGroup):
    waiting_file = State()
    waiting_quantity = State()


@admin_mock_router.message(Command("create_mock_documents"), UserInfo())
async def start_create_mock_documents(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Start mock documents creation process"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return

    await state.set_state(MockDocumentsStates.waiting_file)
    await message.answer("Отправьте файл для тестовых документов")


@admin_mock_router.message(
    StateFilter(MockDocumentsStates.waiting_file), F.photo, UserInfo()
)
async def process_mock_documents_file(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Handle received file and request quantity"""
    await state.update_data(file_id=message.photo[-1].file_id)
    await state.set_state(MockDocumentsStates.waiting_quantity)
    await message.answer("Введите количество документов для каждого объекта (1-5):")


@admin_mock_router.message(
    StateFilter(MockDocumentsStates.waiting_quantity), F.text, UserInfo()
)
async def create_mock_documents(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Create mock documents with specified parameters"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return

    try:
        quantity = int(message.text)
        if not 1 <= quantity <= 5:
            raise ValueError("Quantity out of range")
    except ValueError:
        await message.answer("Пожалуйста, введите число от 1 до 5")
        return

    data = await state.get_data()
    file_id = data.get("file_id")

    async with async_session_maker() as session:
        # Get all active objects
        objects = await ObjectDAO.find_all(
            session, filters=ObjectFilterModel(is_active=True)
        )
        if not objects:
            await message.answer("Нет доступных объектов в базе данных")
            await state.clear()
            return

        # Create documents
        all_documents = []
        doc_types = list(ObjectDocument.DocumentType)
        
        for obj in objects:
            for _ in range(quantity):
                all_documents.append(
                    ObjectDocumentModel(
                        file_id=file_id,
                        object_id=obj.id,
                        document_type=random.choice(doc_types),
                    )
                )

        await ObjectDocumentDAO.add_many(session, all_documents)

    total_docs = len(objects) * quantity
    await message.answer(f"Тестовые документы успешно созданы ({total_docs} шт.)")
    await state.clear()


@admin_mock_router.message(
    StateFilter(MockDocumentsStates.waiting_file), ~F.document, UserInfo()
)
async def process_invalid_document(message: Message, user_info: User) -> None:
    """Handle invalid input in document state"""
    await message.answer("Пожалуйста, отправьте файл документа")

@admin_mock_router.message(Command("create_mock_tools"), UserInfo())
async def start_create_mock_tools(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Start mock tools creation process by requesting photo"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return

    await state.set_state(MockToolsStates.waiting_photo)
    await message.answer("Отправьте фото для тестовых инструментов")


@admin_mock_router.message(
    StateFilter(MockToolsStates.waiting_photo), F.photo, UserInfo()
)
async def process_mock_tools_photo(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Handle received photo and request quantity"""
    await state.update_data(file_id=message.photo[-1].file_id)
    await state.set_state(MockToolsStates.waiting_quantity)
    await message.answer("Введите количество копий каждого инструмента (1-10):")


@admin_mock_router.message(
    StateFilter(MockToolsStates.waiting_quantity), F.text, UserInfo()
)
async def create_mock_tools_with_photo(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Create mock tools with specified photo and quantity"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return

    try:
        quantity = int(message.text)
        if not 1 <= quantity <= 10:
            raise ValueError("Quantity out of range")
    except ValueError:
        await message.answer("Пожалуйста, введите число от 1 до 10")
        return

    data = await state.get_data()
    file_id = data.get("file_id")

    async with async_session_maker() as session:
        for name, description in SAMPLE_TOOLS:
            for _ in range(quantity):
                tool = ToolModel(
                    name=name,
                    description=description,
                    file_id=file_id,
                    status=Tool.Status.in_work,
                    user_id=random.choice(settings.ROOT_ADMIN_IDS),
                )
                await ToolDAO.add(session, tool)

    await message.answer(
        f"Тестовые инструменты успешно созданы ({len(SAMPLE_TOOLS) * quantity} шт.)"
    )
    await state.clear()


@admin_mock_router.message(Command("create_mock_objects"), UserInfo())
async def create_mock_objects(message: Message, user_info: User) -> None:
    """Create mock objects with admins as members"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return
    object_models = []
    async with async_session_maker() as session:
        for name, description in SAMPLE_OBJECTS:
            object_models.append(ObjectModel(
                name=name,
                description=description,
                creator_id=message.from_user.id,
                is_active=True,
            ))
        await ObjectDAO.add_many(session, object_models)
    async with async_session_maker() as session:
        members = []
        objects = await ObjectDAO.find_all(session, filters=ObjectFilterModel(is_active=True))
        for obj in objects:
            for admin_id in settings.ROOT_ADMIN_IDS:
                members.append(ObjectMemberModel(object_id=obj.id, user_id=admin_id))

        await ObjectMemberDAO.add_many(session, members)

    await message.answer(f"Объекты созданны")


@admin_mock_router.message(Command("create_mock_checks"), UserInfo())
async def start_create_mock_checks(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Start mock checks creation process"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return

    await state.set_state(MockObjectChecksStates.waiting_photo)
    await message.answer("Отправьте фото для тестовых чеков")


@admin_mock_router.message(
    StateFilter(MockObjectChecksStates.waiting_photo), F.photo, UserInfo()
)
async def process_mock_checks_photo(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Handle received photo and request quantity"""
    await state.update_data(file_id=message.photo[-1].file_id)
    await state.set_state(MockObjectChecksStates.waiting_quantity)
    await message.answer("Введите количество чеков для каждого объекта (1-10):")


@admin_mock_router.message(
    StateFilter(MockObjectChecksStates.waiting_quantity), F.text, UserInfo()
)
async def create_mock_checks(
    message: Message, state: FSMContext, user_info: User
) -> None:
    """Create mock checks with specified parameters"""
    if message.from_user.id not in settings.ROOT_ADMIN_IDS:
        return

    try:
        quantity = int(message.text)
        if not 1 <= quantity <= 10:
            raise ValueError("Quantity out of range")
    except ValueError:
        await message.answer("Пожалуйста, введите число от 1 до 10")
        return

    data = await state.get_data()
    file_id = data.get("file_id")
    
    # First get all objects and their members
    object_data = []
    async with async_session_maker() as session:
        objects = await ObjectDAO.find_all(
            session, 
            filters=ObjectFilterModel(is_active=True)
        )
        if not objects:
            await message.answer("Нет доступных объектов в базе данных")
            await state.clear()
            return
            
        for obj in objects:
            members = await ObjectMemberDAO.find_object_members(session, obj.id)
            if members:
                object_data.append((obj, members))

    # Then create and save all checks
    if object_data:
        all_checks = []
        for obj, members in object_data:
            for _ in range(quantity):
                amount = round(random.uniform(1000, 100000), 2)
                all_checks.append(
                    ObjectCheckModel(
                        file_id=file_id,
                        amount=amount,
                        description=f"Тестовый чек для объекта {obj.name}",
                        own_expense=random.choice([True, False]),
                        object_id=obj.id,
                        user_id=random.choice(members).telegram_id,
                    )
                )
        
        # Save all checks in a single transaction
        async with async_session_maker() as session:
            await ObjectCheckDAO.add_many(session, all_checks)
            
        total_checks = len(object_data) * quantity
        await message.answer(f"Тестовые чеки успешно созданы ({total_checks} шт.)")
    else:
        await message.answer("Нет объектов с прикрепленными пользователями")
    
    await state.clear()


from openpyxl import Workbook
import io

def generate_test_tmc_data() -> io.BytesIO:
    """Generate test TMC data file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон инструментов"

    # Headers
    headers = [
        "Наименование*",
        "Количество*",
        "Описание",
        "Статус",
        "File ID фото"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Test data
    test_data = [
        ["Перфоратор Bosch GBH 2-26", "2", "Новый, в комплекте 2 бура", "свободный", None],
        ["Шуруповерт Makita", "3", "Аккумулятор 4А", "свободный", None],
        ["Болгарка DeWalt", "1", "180мм", "свободный", None],
        ["Уровень лазерный Bosch", "2", "Зеленый луч", "свободный", None],
        # Test invalid data
        ["", "2", "Пустое имя", "свободный", None],  # Should be skipped
        ["Дрель", "-1", "Отрицательное количество", "свободный", None],  # Should be skipped
        ["Отвертка", "abc", "Неверное количество", "свободный", None],  # Should be skipped
    ]

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

@admin_mock_router.message(F.text == '/test_tmc_file', UserInfo())
async def process_tmc_test_file(message:Message, user_info: User):
    """Send test TMC file for testing"""

    test_file = generate_test_tmc_data()
    await message.answer_document(
        document=BufferedInputFile(
            test_file.getvalue(),
            filename="tmc_test_data.xlsx"
        ),
        caption=get_text("tmc_test_file_instruction", user_info.language)
    )
